"""
Konvertiert die Modulsteckbrief-Arbeitsmappe in eine saubere modules.json.

Hintergrund (Masterarbeit):
Die Zielgruppen stehen in den Steckbriefen als Fließtext. Für eine belastbare,
filterbare Datenbasis werden sie zusätzlich auf die kontrollierte Zielgruppen-
Liste aus dem Blatt LISTEN normiert (kontrolliertes Vokabular). Das Freitext-
Original bleibt erhalten, sodass keine Information verloren geht.

Nutzung:
    python scripts/convert_xlsx_to_json.py \
        --input data/Arbeitsergebnis_01_Modulsteckbriefe.xlsx \
        --output data/modules.json
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl

# --- Reihenfolge der Steckbrief-Abschnitte (für die volle Anzeige) -----------
SECTION_HEADERS = [
    "1  Identifikation",
    "2  Zielgruppe & Adressierung",
    "3  Lernziele & Kompetenzen",
    "4  Inhalt & Prozessbezug (Remanufacturing)",
    "5  Didaktik & Methodik",
    "6  Organisation & Einordnung",
    "7  Evaluation & Erfolgskriterien",
    "8  Wirtschaftlichkeit & Trägerschaft",
]

# --- Kontrolliertes Zielgruppen-Vokabular (aus Blatt LISTEN) ------------------
# Schlüssel = normierte Kategorie; Werte = Erkennungsmuster (Kleinschreibung,
# Teilstring-Match) im Steckbrief-Freitext. Bewusst konservativ gehalten.
ZIELGRUPPEN_KANON = {
    "Schüler:innen": ["schüler"],
    "Bachelor": ["bachelor"],
    "Master": ["master", "graduate campus"],
    "Promovierende": ["promov", "forschende", "phd", "doktorand"],
    "Auszubildende": ["auszubild", "azubi"],
    "Fachkräfte": ["fachkräfte", "fachkraft"],
    "Führungskräfte": ["führungskräfte", "führungskraft", "management", "manager"],
    "Unternehmer:innen": ["unternehmer", "industrie", "projektteam"],
    "Freiberufler:innen": ["freiberuf"],
    "Gewerkschaft": ["gewerkschaft"],
    "Öffentlichkeit": ["öffentlichkeit", "messepublikum", "messe", "public"],
}


def is_module_sheet(name: str) -> bool:
    """Modul-Blätter beginnen mit zwei Ziffern, z. B. '01 RecycleBot ...'."""
    return bool(re.match(r"^\d{2}\s", name))


def normiere_zielgruppen(freitext: str) -> list[str]:
    """Mappt den Zielgruppen-Freitext auf kontrollierte Kategorien."""
    if not freitext:
        return []
    low = freitext.lower()
    treffer = []
    for kategorie, muster in ZIELGRUPPEN_KANON.items():
        if any(m in low for m in muster):
            treffer.append(kategorie)
    return treffer


def lies_modulblatt(ws) -> dict:
    """Liest ein Key-Value-Steckbriefblatt in ein strukturiertes Dict."""
    felder: dict[str, str] = {}
    abschnitte: dict[str, list[tuple[str, str]]] = {}
    aktueller_abschnitt = "Allgemein"

    for row in ws.iter_rows(values_only=True):
        a = (row[0] if len(row) > 0 else None)
        b = (row[1] if len(row) > 1 else None)
        if a is None:
            continue
        a = str(a).strip()

        # Abschnitts-Überschrift erkennen (Spalte B leer, Text matcht Header)
        if b is None and a in SECTION_HEADERS:
            aktueller_abschnitt = a
            abschnitte.setdefault(aktueller_abschnitt, [])
            continue

        if b is not None:
            wert = str(b).strip()
            felder[a] = wert
            abschnitte.setdefault(aktueller_abschnitt, []).append((a, wert))

    zielgruppe_text = felder.get("Zielgruppe(n)", "")

    return {
        "modul_id": felder.get("Modul-ID / Kürzel", ""),
        "name": felder.get("Modulname", ""),
        "version": felder.get("Version / Stand", ""),
        "autor": felder.get("Autor:in / verantwortlich", ""),
        "zielgruppe_text": zielgruppe_text,
        "zielgruppe_kanon": normiere_zielgruppen(zielgruppe_text),
        "lernziel": felder.get("Übergeordnetes Lernziel", ""),
        "kompetenzklassen": felder.get("Kompetenzklassen (Erpenbeck)", ""),
        "bloom_stufen": felder.get("Kognitive Stufen (Bloom)", ""),
        "hauptzweck": felder.get("Hauptzweck", ""),
        "dauer": felder.get("Dauer", ""),
        "status": felder.get("Status / Reifegrad", ""),
        # vollständige Felder + Abschnittsstruktur für die Detailansicht
        "felder": felder,
        "abschnitte": {
            sek: [{"label": k, "wert": v} for k, v in paare]
            for sek, paare in abschnitte.items()
            if paare
        },
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    module = []
    for name in wb.sheetnames:
        if is_module_sheet(name):
            m = lies_modulblatt(wb[name])
            m["blatt"] = name
            module.append(m)

    # alle in der Datenbasis tatsächlich vorkommenden Kanon-Zielgruppen
    vorhandene_zg = sorted(
        {z for m in module for z in m["zielgruppe_kanon"]},
        key=lambda x: list(ZIELGRUPPEN_KANON).index(x),
    )

    ausgabe = {
        "meta": {
            "quelle": Path(args.input).name,
            "anzahl_module": len(module),
            "zielgruppen_kanon_alle": list(ZIELGRUPPEN_KANON),
            "zielgruppen_kanon_vorhanden": vorhandene_zg,
            "abschnitte": SECTION_HEADERS,
        },
        "module": module,
    }

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(ausgabe, f, ensure_ascii=False, indent=2)

    print(f"{len(module)} Module -> {args.output}")
    print("Vorhandene Zielgruppen:", ", ".join(vorhandene_zg))


if __name__ == "__main__":
    main()
