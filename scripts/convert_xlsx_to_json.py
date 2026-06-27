"""
Konvertiert die Modulsteckbrief-Arbeitsmappe in eine saubere modules.json.

Hintergrund (Masterarbeit):
Die Zielgruppen stehen in den Steckbriefen als Fließtext. Für eine belastbare,
filterbare Datenbasis werden sie zusätzlich auf die kontrollierte Zielgruppen-
Liste aus dem Blatt LISTEN normiert (kontrolliertes Vokabular). Das Freitext-
Original bleibt erhalten, sodass keine Information verloren geht.

Nutzung:
    python scripts/convert_xlsx_to_json.py \
        --input data/Arbeitsergebnis_01_Modulsteckbriefe.xlsx \
        --output data/modules.json
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl

# --- Reihenfolge der Steckbrief-Abschnitte (für die volle Anzeige) -----------
SECTION_HEADERS = [
    "1  Identifikation",
    "2  Zielgruppe & Adressierung",
    "3  Lernziele & Kompetenzen",
    "4  Inhalt & Prozessbezug (Remanufacturing)",
    "5  Didaktik & Methodik",
    "6  Organisation & Einordnung",
    "7  Evaluation & Erfolgskriterien",
    "8  Wirtschaftlichkeit & Trägerschaft",
]

# --- Kontrolliertes Zielgruppen-Vokabular (aus Blatt LISTEN) ------------------
# Schlüssel = normierte Kategorie; Werte = Erkennungsmuster (Kleinschreibung,
# Teilstring-Match) im Steckbrief-Freitext. Bewusst konservativ gehalten.
ZIELGRUPPEN_KANON = {
    "Schüler:innen": ["schüler"],
    "Bachelor": ["bachelor"],
    "Master": ["master", "graduate campus"],
    "Promovierende": ["promov", "forschende", "phd", "doktorand"],
    "Auszubildende": ["auszubild", "azubi"],
    "Fachkräfte": ["fachkräfte", "fachkraft"],
    "Führungskräfte": ["führungskräfte", "führungskraft", "management", "manager"],
    "Unternehmer:innen": ["unternehmer", "industrie", "projektteam"],
    "Freiberufler:innen": ["freiberuf"],
    "Gewerkschaft": ["gewerkschaft"],
    "Öffentlichkeit": ["öffentlichkeit", "messepublikum", "messe", "public"],
}


def is_module_sheet(name: str) -> bool:
    """Modul-Blätter beginnen mit zwei Ziffern, z. B. '01 RecycleBot ...'."""
    return bool(re.match(r"^\d{2}\s", name))


def normiere_zielgruppen(freitext: str) -> list[str]:
    """Mappt den Zielgruppen-Freitext auf kontrollierte Kategorien."""
    if not freitext:
        return []
    low = freitext.lower()
    treffer = []
    for kategorie, muster in ZIELGRUPPEN_KANON.items():
        if any(m in low for m in muster):
            treffer.append(kategorie)
    return treffer


def lies_modulblatt(ws) -> dict:
    """Liest ein Key-Value-Steckbriefblatt in ein strukturiertes Dict."""
    felder: dict[str, str] = {}
    abschnitte: dict[str, list[tuple[str, str]]] = {}
    aktueller_abschnitt = "Allgemein"

    for row in ws.iter_rows(values_only=True):
        a = (row[0] if len(row) > 0 else None)
        b = (row[1] if len(row) > 1 else None)
        if a is None:
            continue
        a = str(a).strip()

        # Abschnitts-Überschrift erkennen (Spalte B leer, Text matcht Header)
        if b is None and a in SECTION_HEADERS:
            aktueller_abschnitt = a
            abschnitte.setdefault(aktueller_abschnitt, [])
            continue

        if b is not None:
            wert = str(b).strip()
            felder[a] = wert
            abschnitte.setdefault(aktueller_abschnitt, []).append((a, wert))

    zielgruppe_text = felder.get("Zielgruppe(n)", "")

    return {
        "modul_id": felder.get("Modul-ID / Kürzel", ""),
        "name": felder.get("Modulname", ""),
        "version": felder.get("Version / Stand", ""),
        "autor": felder.get("Autor:in / verantwortlich", ""),
        "zielgruppe_text": zielgruppe_text,
        "zielgruppe_kanon": normiere_zielgruppen(zielgruppe_text),
        "lernziel": felder.get("Übergeordnetes Lernziel", ""),
        "kompetenzklassen": felder.get("Kompetenzklassen (Erpenbeck)", ""),
        "bloom_stufen": felder.get("Kognitive Stufen (Bloom)", ""),
        "hauptzweck": felder.get("Hauptzweck", ""),
        "dauer": felder.get("Dauer", ""),
        "status": felder.get("Status / Reifegrad", ""),
        # vollständige Felder + Abschnittsstruktur für die Detailansicht
        "felder": felder,
        "abschnitte": {
            sek: [{"label": k, "wert": v} for k, v in paare]
            for sek, paare in abschnitte.items()
            if paare
        },
    }


def lies_uebersicht(wb) -> dict:
    """Liest die Schwarzer-Übersicht (Level, Zustand, Ansprechpartner) je Modulname."""
    blatt = "Learning Factory Module Overvie"
    if blatt not in wb.sheetnames:
        return {}
    ws = wb[blatt]
    rows = list(ws.iter_rows(values_only=True))
    info = {}
    for r in rows[2:]:  # erste zwei Zeilen sind Kopf
        titel = (r[0] if len(r) > 0 else None)
        if not titel:
            continue
        info[str(titel).strip()] = {
            "kurzbeschreibung": str(r[1]).strip() if len(r) > 1 and r[1] else "",
            "level": str(r[3]).strip() if len(r) > 3 and r[3] else "",
            "zustand": str(r[4]).strip() if len(r) > 4 and r[4] else "",
            "ansprechpartner": str(r[5]).strip() if len(r) > 5 and r[5] else "",
        }
    return info


def lies_schwarzer_steckbrief(wb) -> dict | None:
    """Liest das 3-spaltige Steckbrief-Blatt von Prof. Schwarzer."""
    blatt = "Modulsteckbrief_Schwarzer"
    if blatt not in wb.sheetnames:
        return None
    ws = wb[blatt]
    felder, abschnitte = {}, {}
    aktueller = "Allgemein"
    for r in ws.iter_rows(values_only=True):
        a = (r[0] if len(r) > 0 else None)
        b = (r[1] if len(r) > 1 else None)
        if a is None:
            continue
        a = str(a).strip()
        # Schwarzer nutzt leicht andere Abschnittstitel (z.B. "4  Inhalt & Prozessbezug")
        if (b is None or str(b).strip() == "") and a in SECTION_HEADERS_SCHWARZER:
            aktueller = SCHWARZER_ZU_KANON.get(a, a)
            abschnitte.setdefault(aktueller, [])
            continue
        if b is not None and str(b).strip():
            wert = str(b).strip()
            felder[a] = wert
            abschnitte.setdefault(aktueller, []).append((a, wert))
    zg = felder.get("Zielgruppe(n)", "")
    return {
        "blatt": "Modulsteckbrief_Schwarzer",
        "modul_id": felder.get("Modul-ID / Kürzel", ""),
        "name": felder.get("Modulname", ""),
        "version": felder.get("Version / Stand", ""),
        "autor": felder.get("Autor:in / verantwortlich", ""),
        "zielgruppe_text": zg,
        "zielgruppe_kanon": normiere_zielgruppen(zg),
        "lernziel": felder.get("Übergeordnetes Lernziel", ""),
        "kompetenzklassen": felder.get("Kompetenzklassen (Erpenbeck)", ""),
        "bloom_stufen": felder.get("Kognitive Stufen (Bloom)", ""),
        "hauptzweck": felder.get("Hauptzweck", ""),
        "dauer": felder.get("Dauer", ""),
        "status": felder.get("Status / Reifegrad", "")
        or felder.get("Strategische Relevanz", ""),
        "quelle_datei": "Arbeitsergebnis_02_Ergaenzung_Schwarzer.xlsx",
        "felder": felder,
        "abschnitte": {
            sek: [{"label": k, "wert": v} for k, v in paare]
            for sek, paare in abschnitte.items()
            if paare
        },
    }


# Schwarzer verwendet teils verkürzte Abschnittsüberschriften
SECTION_HEADERS_SCHWARZER = [
    "1  Identifikation",
    "2  Zielgruppe & Adressierung",
    "3  Lernziele & Kompetenzen",
    "4  Inhalt & Prozessbezug",
    "5  Didaktik & Methodik",
    "6  Organisation & Einordnung",
    "7  Evaluation & Erfolgskriterien",
    "8  Wirtschaftlichkeit & Trägerschaft",
]
SCHWARZER_ZU_KANON = {
    "4  Inhalt & Prozessbezug": "4  Inhalt & Prozessbezug (Remanufacturing)",
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True,
                    help="Haupt-Arbeitsmappe (25 Module)")
    ap.add_argument("--schwarzer", default=None,
                    help="Optional: Schwarzer-Ergänzungsmappe")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    module = []
    for name in wb.sheetnames:
        if is_module_sheet(name):
            m = lies_modulblatt(wb[name])
            m["blatt"] = name
            m["quelle_datei"] = Path(args.input).name
            module.append(m)

    # --- Schwarzer-Ergänzung: Übersichtsfelder + 26. Modul -------------------
    if args.schwarzer:
        wb2 = openpyxl.load_workbook(args.schwarzer, read_only=True, data_only=True)
        uebersicht = lies_uebersicht(wb2)
        # Level/Zustand/Ansprechpartner an bestehende Module anreichern
        for m in module:
            info = uebersicht.get(m["name"])
            if info:
                m["level"] = info["level"]
                m["zustand"] = info["zustand"]
                m["ansprechpartner"] = info["ansprechpartner"]
        # Hinweis: Schwarzers eigenes Modul wird bewusst NICHT automatisch
        # eingetragen – es soll in der App über das Eingabeformular erfasst
        # werden. (lies_schwarzer_steckbrief bleibt als optionale Hilfe erhalten.)

    # Felder mit Defaults absichern
    for m in module:
        m.setdefault("level", "")
        m.setdefault("zustand", "")
        m.setdefault("ansprechpartner", "")
        m.setdefault("manuell", False)

    # alle in der Datenbasis tatsächlich vorkommenden Kanon-Zielgruppen
    vorhandene_zg = sorted(
        {z for m in module for z in m["zielgruppe_kanon"]},
        key=lambda x: list(ZIELGRUPPEN_KANON).index(x),
    )

    # LISTEN (kontrolliertes Vokabular) mitliefern, falls vorhanden
    listen = {}
    listen_pfad = Path(args.output).parent / "listen.json"
    if listen_pfad.exists():
        with open(listen_pfad, encoding="utf-8") as f:
            listen = json.load(f)

    ausgabe = {
        "meta": {
            "quelle": Path(args.input).name,
            "anzahl_module": len(module),
            "zielgruppen_kanon_alle": list(ZIELGRUPPEN_KANON),
            "zielgruppen_kanon_vorhanden": vorhandene_zg,
            "abschnitte": SECTION_HEADERS,
            "listen": listen,
        },
        "module": module,
    }

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(ausgabe, f, ensure_ascii=False, indent=2)

    print(f"{len(module)} Module -> {args.output}")
    print("Vorhandene Zielgruppen:", ", ".join(vorhandene_zg))


if __name__ == "__main__":
    main()
